from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, event, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.database import get_db
from app.main import app
from app.models import Base, LanguagePreference, Part, PartAlias, SysConfig
from app.services.excel_batch import parse_excel

engine = create_engine("sqlite+pysqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)


@event.listens_for(engine, "connect")
def foreign_keys(connection, _record):
    connection.execute("PRAGMA foreign_keys=ON")


TestingSession = sessionmaker(bind=engine, expire_on_commit=False)
Base.metadata.create_all(engine)
client = TestClient(app, client=("127.0.0.1", 51000))
A = {"X-Session-Id": "language-owner-a"}
B = {"X-Session-Id": "language-owner-b"}


def override_db():
    with TestingSession() as session:
        yield session


def body(response, status=200):
    assert response.status_code == status, response.text
    payload = response.json()
    assert set(payload) == {"code", "message", "data"}
    assert payload["code"] == 0
    return payload["data"]


@pytest.fixture(autouse=True)
def catalogue():
    old = app.dependency_overrides.get(get_db)
    app.dependency_overrides[get_db] = override_db
    client.cookies.clear()
    with TestingSession.begin() as db:
        for table in reversed(Base.metadata.sorted_tables):
            db.execute(table.delete())
        primary = Part(sku="I18N-1", part_no="I-100", oem_no="IOEM-1", brand="Acme", category="filters",
                       name_zh="空气滤芯", name_en="Air Filter", name_vi="Lọc gió", specs={}, stock=3, is_active=True)
        fallback = Part(sku="I18N-2", part_no="I-200", oem_no="IOEM-2", brand="Acme", category="filters",
                        name_zh="机油滤芯", name_en="Oil Filter", name_vi=None, specs={}, stock=2, is_active=True)
        db.add_all([primary, fallback]); db.flush()
        db.add_all([PartAlias(part_id=primary.id, alias="air cleaner", language="en", status="active"),
                    PartAlias(part_id=primary.id, alias="lọc khí", language="vi", status="active")])
    yield {"primary": primary.id, "fallback": fallback.id}
    if old is None:
        app.dependency_overrides.pop(get_db, None)
    else:
        app.dependency_overrides[get_db] = old


def test_t1_languages_messages_envelopes_defaults_and_sys_config_override():
    languages = body(client.get("/api/v1/i18n/languages", headers={"Accept-Language": "vi-VN, en;q=0.8"}))
    assert languages["current"] == "vi"
    assert [item["code"] for item in languages["languages"]] == ["zh", "en", "vi"]
    english = body(client.get("/api/v1/i18n/messages?lang=en"))
    assert english["lang"] == "en" and english["messages"]["common.ok"] == "Success"
    with TestingSession.begin() as db:
        db.add(SysConfig(key="i18n.messages", value={"vi": {"common.ok": "Đã xong"}}))
    vietnamese = body(client.get("/api/v1/i18n/messages?lang=vi"))
    assert vietnamese["messages"]["common.ok"] == "Đã xong"
    assert vietnamese["messages"]["error.not_found"]  # defaults survive partial overrides


def test_t2_names_localize_for_exact_text_detail_hot_and_fallback(catalogue):
    exact = body(client.get("/api/v1/search?type=part_no&q=I-100&lang=vi"))
    assert exact["candidates"][0]["part"]["name"] == "Lọc gió"
    fallback = body(client.get("/api/v1/search?type=part_no&q=I-200&lang=vi"))
    assert fallback["candidates"][0]["part"]["name"] == "Oil Filter"
    detail = body(client.get(f"/api/v1/parts/{catalogue['primary']}?lang=en"))
    assert detail["name"] == "Air Filter"
    hot = body(client.get("/api/v1/parts/hot?lang=zh"))
    assert {item["name"] for item in hot} == {"空气滤芯", "机油滤芯"}


def test_t3_preference_persists_is_owner_isolated_and_sets_hardened_cookie(catalogue):
    response = client.post("/api/v1/i18n/preference", headers=A, json={"lang": "vi"})
    body(response)
    cookie = response.headers["set-cookie"].lower()
    assert "httponly" in cookie and "samesite=lax" in cookie and "secure" in cookie
    persisted = body(client.get("/api/v1/search?type=part_no&q=I-100", headers=A))
    assert persisted["candidates"][0]["part"]["name"] == "Lọc gió"
    isolated = body(client.get("/api/v1/search?type=part_no&q=I-100", headers=B | {"Accept-Language": "en"}))
    assert isolated["candidates"][0]["part"]["name"] == "Air Filter"
    explicit = body(client.get("/api/v1/search?type=part_no&q=I-100&lang=zh", headers=A))
    assert explicit["candidates"][0]["part"]["name"] == "空气滤芯"
    with TestingSession() as db:
        stored = db.scalar(select(LanguagePreference).where(LanguagePreference.owner_key == "session:language-owner-a"))
        assert stored.language == "vi" and stored.session_id == "language-owner-a"


def test_t3_forged_forwarding_and_country_headers_are_ignored_until_peer_is_trusted():
    forged = body(client.get("/api/v1/i18n/languages", headers={
        "X-Forwarded-For": "1.2.3.4", "CF-IPCountry": "CN", "Accept-Language": "vi",
    }))
    assert forged["current"] == "vi"
    with TestingSession.begin() as db:
        db.add(SysConfig(key="i18n.trusted_proxy_ips", value=["127.0.0.1/32"]))
        db.add(SysConfig(key="i18n.country_header", value="CF-IPCountry"))
    trusted = body(client.get("/api/v1/i18n/languages", headers={"CF-IPCountry": "CN", "Accept-Language": "vi"}))
    assert trusted["current"] == "zh"


def test_t4_alias_language_priority_with_controlled_en_zh_fallback(catalogue):
    vietnamese = body(client.get("/api/v1/search?type=text&q=lọc%20khí&lang=vi"))
    assert vietnamese["candidates"][0]["part"]["id"] == catalogue["primary"]
    assert vietnamese["candidates"][0]["confidence"] == 1
    english_fallback = body(client.get("/api/v1/search?type=text&q=air%20cleaner&lang=vi"))
    assert english_fallback["candidates"][0]["part"]["id"] == catalogue["primary"]
    assert english_fallback["candidates"][0]["confidence"] == .9


@pytest.mark.parametrize("lang,first_header,sheet_name", [
    ("zh", "设备类型", "填写说明"), ("en", "Machine Type", "Instructions"),
    ("vi", "Loại thiết bị", "Hướng dẫn"),
])
def test_t5_templates_are_translated_and_round_trip_through_upload_parser(lang, first_header, sheet_name):
    response = client.get(f"/api/v1/batch/template?lang={lang}")
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.active["A1"].value == first_header and sheet_name in workbook.sheetnames
    assert workbook.active["K1"].comment and workbook.active["K2"].value == 2
    parsed, errors = parse_excel(".xlsx", response.content, 500)
    assert not errors and parsed[0]["quantity"] == 2 and parsed[0]["normalized"]["part_no"] == "12345-67890"


def test_postgres_incremental_migration_contains_normalized_table_and_sys_configs():
    migration = (Path(__file__).parents[1] / "migrations/versions/f19b2c4d8e60_i18n_preferences.py").read_text()
    assert 'down_revision: Union[str, None] = "e7a6c4b2d901"' in migration
    assert '"language_preference"' in migration and 'sa.UniqueConstraint("owner_key"' in migration
    assert '"i18n.trusted_proxy_ips"' in migration and '"i18n.cookie_secure"' in migration
