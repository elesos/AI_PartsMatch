from __future__ import annotations

from io import BytesIO
from time import perf_counter

import pytest
import xlwt
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, event, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.database import get_db
from app.main import app
from app.models import (Base, CartItem, ExcelBatch, ExcelBatchJob, ExcelBatchRow, FileObject,
                        ManualTicket, Part, PartQueryLog)
from app.services.excel_batch import create_template, parse_excel, validate_excel_file
from app.services.storage import StorageService

engine = create_engine("sqlite+pysqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)


@event.listens_for(engine, "connect")
def foreign_keys(connection, _record):
    connection.execute("PRAGMA foreign_keys=ON")


TestingSession = sessionmaker(bind=engine, expire_on_commit=False)
Base.metadata.create_all(engine)
client = TestClient(app)
A = {"X-Session-Id": "excel-session-owner-a"}
B = {"X-Session-Id": "excel-session-owner-b"}


def override_db():
    with TestingSession() as session:
        yield session


def body(response, status=200):
    assert response.status_code == status, response.text
    payload = response.json()
    assert set(payload) == {"code", "message", "data"}
    return payload["data"]


@pytest.fixture(autouse=True)
def database(monkeypatch):
    old = app.dependency_overrides.get(get_db)
    app.dependency_overrides[get_db] = override_db
    with TestingSession.begin() as db:
        for table in reversed(Base.metadata.sorted_tables):
            db.execute(table.delete())
        db.add(Part(sku="BATCH-1", part_no="P-100", oem_no="OEM-100", brand="Acme", category="filters",
                    name_zh="机油滤芯", name_en="Oil filter", specs={}, stock=5, is_active=True))

    async def fake_upload(self, upload, *, owner_key=None, images_only=False):
        content = await upload.read()
        record = FileObject(object_key=f"tests/{owner_key}/{upload.filename}", original_name=upload.filename,
                            mime_type=upload.content_type, size=len(content), url="https://files.test/input",
                            owner_key=owner_key)
        self.db.add(record)
        self.db.commit()
        self.db.refresh(record)
        return record

    monkeypatch.setattr(StorageService, "upload", fake_upload)
    yield
    if old is None:
        app.dependency_overrides.pop(get_db, None)
    else:
        app.dependency_overrides[get_db] = old


def xlsx(rows, headers=None):
    wb = Workbook()
    ws = wb.active
    ws.append(headers or ["Part Number", "所需数量", "备注"])
    for row in rows:
        ws.append(row)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def xls(rows, headers=None):
    book = xlwt.Workbook()
    sheet = book.add_sheet("批量匹配")
    values = headers or ["Part Number", "Quantity", "Notes"]
    for col, value in enumerate(values):
        sheet.write(0, col, value)
    for row_index, row in enumerate(rows, 1):
        for col, value in enumerate(row):
            sheet.write(row_index, col, value)
    stream = BytesIO()
    book.save(stream)
    return stream.getvalue()


def upload(content, *, name="parts.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=A):
    return client.post("/api/v1/batch/upload", headers=headers, files={"file": (name, content, mime)})


def test_t1_template_is_real_xlsx_with_instructions_comments_and_example():
    response = client.get("/api/v1/batch/template")
    assert response.status_code == 200 and response.content.startswith(b"PK\x03\x04")
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["批量匹配", "填写说明"]
    sheet = workbook["批量匹配"]
    assert sheet["A1"].value == "设备类型" and sheet["K1"].value == "所需数量"
    assert sheet["K1"].comment and sheet["A2"].value == "叉车" and sheet["K2"].value == 2


def test_t2_real_xlsx_and_xls_bilingual_parsing_validation_and_duplicates():
    uploaded = body(upload(xlsx([["P-100", 2, "ok"], ["P-100", 2, "duplicate"], ["", 0, "bad"]])), 201)
    assert uploaded["total_rows"] == 3 and uploaded["valid_rows"] == 2
    assert uploaded["validation_errors"][0]["row_index"] == 4
    assert uploaded["duplicate_rows"] == [{"part_number": "p-100", "quantity": 2,
                                            "row_indexes": [2, 3], "suggestion": "merge"}]
    with TestingSession() as db:
        log = db.scalar(select(PartQueryLog).where(PartQueryLog.query_type == "excel"))
        assert log is not None and log.source_id == uploaded["batch_id"] and log.request_data["file_id"] == uploaded["file_id"]
    parsed, errors = parse_excel(".xls", xls([["P-100", 1, "legacy"]]), 500)
    assert parsed[0]["normalized"]["part_no"] == "P-100" and parsed[0]["quantity"] == 1 and not errors


def test_t2_rejects_extension_mime_magic_formula_limits_and_zip_bomb():
    assert upload(b"not excel").status_code == 415
    assert upload(xlsx([["P-100", 1, ""]]), mime="application/vnd.ms-excel").status_code == 415
    formula = xlsx([["P-100", 1, "=HYPERLINK(\"https://evil\")"]])
    response = upload(formula)
    assert response.status_code == 400 and response.json()["code"] == "EXCEL_FORMULA_REJECTED"
    too_many = upload(xlsx([[f"P-{index}", 1, ""] for index in range(501)]))
    assert too_many.status_code == 400 and too_many.json()["code"] == "BATCH_TOO_MANY_ROWS"
    with TestingSession.begin() as db:
        from app.models import SysConfig
        db.add(SysConfig(key="batch.max_zip_ratio", value=1))
    bomb = upload(xlsx([["P-100", 1, "A" * 10000]]))
    assert bomb.status_code == 413 and bomb.json()["code"] == "EXCEL_ZIP_BOMB"


def test_t2_owner_isolation_and_xls_formula_rejection():
    batch_id = body(upload(xlsx([["P-100", 1, ""]])), 201)["batch_id"]
    assert client.get(f"/api/v1/batch/{batch_id}", headers=B).status_code == 404
    formula_book = xlwt.Workbook()
    sheet = formula_book.add_sheet("Sheet1")
    for col, value in enumerate(["Part Number", "Quantity", "Notes"]):
        sheet.write(0, col, value)
    sheet.write(1, 0, "P-100"); sheet.write(1, 1, 1); sheet.write(1, 2, xlwt.Formula("1+1"))
    stream = BytesIO(); formula_book.save(stream)
    response = upload(stream.getvalue(), name="legacy.xls", mime="application/vnd.ms-excel")
    assert response.status_code == 400 and response.json()["code"] == "EXCEL_FORMULA_REJECTED"


def test_t3_t5_t6_matching_atomic_cart_and_idempotent_tickets():
    batch_id = body(upload(xlsx([["P-100", 2, ""], ["UNKNOWN", 3, ""]])), 201)["batch_id"]
    matched = body(client.post(f"/api/v1/batch/{batch_id}/match", headers=A))
    assert matched["mode"] == "sync"
    exact, missing = matched["rows"]
    assert exact["match_status"] == "exact" and exact["suggested_action"] == "confirm"
    assert missing["match_status"] == "not_found" and missing["suggested_action"] == "manual"
    part_id = exact["candidates"][0]["part"]["id"]
    rejected = client.post(f"/api/v1/batch/{batch_id}/add-to-cart", headers=A, json={"selections": [
        {"row_index": 2, "part_id": part_id, "quantity": 2},
        {"row_index": 3, "part_id": part_id, "quantity": 3, "confirmed": True},
    ]})
    assert rejected.status_code == 409 and rejected.json()["data"]["atomic"] is True
    with TestingSession() as db:
        assert db.scalar(select(CartItem)) is None
    added = body(client.post(f"/api/v1/batch/{batch_id}/add-to-cart", headers=A, json={"selections": [
        {"row_index": 2, "part_id": part_id, "quantity": 2}
    ]}))
    assert added["atomic"] is True and added["added"][0]["quantity"] == 2
    request = {"row_indexes": [3], "contact_name": "Buyer", "contact_info": "+8613800000000",
               "communication_tool": "wechat"}
    first = body(client.post(f"/api/v1/batch/{batch_id}/create-tickets", headers=A, json=request))
    second = body(client.post(f"/api/v1/batch/{batch_id}/create-tickets", headers=A, json=request))
    assert len(first["created"]) == 1 and second["existing"] == first["created"]
    with TestingSession() as db:
        ticket = db.get(ManualTicket, first["created"][0]["ticket_id"])
        row = db.scalar(select(ExcelBatchRow).where(ExcelBatchRow.batch_id == batch_id,
                                                    ExcelBatchRow.row_index == 3))
        assert ticket.excel_batch_id == batch_id and "原始行 3" in ticket.note and row.ticket_id == ticket.id


def test_t3_100_rows_use_persisted_job_polling_and_finish_under_30_seconds():
    content = xlsx([["P-100", 1, str(index)] for index in range(100)])
    batch_id = body(upload(content), 201)["batch_id"]
    started = perf_counter()
    response = body(client.post(f"/api/v1/batch/{batch_id}/match", headers=A))
    elapsed = perf_counter() - started
    assert response["mode"] == "async" and response["job_id"]
    job = body(client.get(f"/api/v1/batch/jobs/{response['job_id']}", headers=A))
    assert job["status"] == "completed" and job["processed_rows"] == 100 and job["attempts"] == 1
    status = body(client.get(f"/api/v1/batch/{batch_id}/status", headers=A))
    assert status["job_id"] == response["job_id"] and status["processed_rows"] == status["total_rows"] == 100
    assert elapsed < 30
    with TestingSession() as db:
        persisted = db.get(ExcelBatchJob, response["job_id"])
        assert persisted.status == "completed" and db.get(ExcelBatch, batch_id).status == "matched"
    assert client.get(f"/api/v1/batch/jobs/{response['job_id']}", headers=B).status_code == 404
    assert client.get(f"/api/v1/batch/{batch_id}/status", headers=B).status_code == 404


def test_failed_job_is_traceable_and_retryable(monkeypatch):
    batch_id = body(upload(xlsx([["P-100", 1, ""] for _ in range(51)])), 201)["batch_id"]
    from app.services.excel_batch import ExcelBatchService
    original = ExcelBatchService.match_all

    def fail(*args, **kwargs):
        raise RuntimeError("worker exploded")

    monkeypatch.setattr(ExcelBatchService, "match_all", fail)
    response = body(client.post(f"/api/v1/batch/{batch_id}/match", headers=A))
    failed = body(client.get(f"/api/v1/batch/jobs/{response['job_id']}", headers=A))
    assert failed["status"] == "failed" and "worker exploded" in failed["error"]
    monkeypatch.setattr(ExcelBatchService, "match_all", original)
    retried = body(client.post(f"/api/v1/batch/jobs/{response['job_id']}/retry", headers=A))
    assert retried["status"] == "retrying"
    completed = body(client.get(f"/api/v1/batch/jobs/{response['job_id']}", headers=A))
    assert completed["status"] == "completed" and completed["attempts"] == 2


def test_patch_row_updates_allowed_fields_revalidates_rematches_and_checks_owner():
    batch_id = body(upload(xlsx([["UNKNOWN", 2, ""]])), 201)["batch_id"]
    body(client.post(f"/api/v1/batch/{batch_id}/match", headers=A))
    updated = body(client.patch(f"/api/v1/batch/{batch_id}/rows/2", headers=A, json={
        "machine_brand": "Acme", "model": "M-20", "part_no": "P-100", "system": "Hydraulic",
        "quantity": 4,
    }))
    assert updated["match_status"] == "exact" and updated["quantity"] == 4
    assert updated["normalized_content"]["machine_model"] == "M-20"
    assert updated["normalized_content"]["part_system"] == "Hydraulic"
    assert client.patch(f"/api/v1/batch/{batch_id}/rows/2", headers=B, json={"part_no": "P-100"}).status_code == 404
    rejected = client.patch(f"/api/v1/batch/{batch_id}/rows/2", headers=A, json={"serial_no": "forbidden"})
    assert rejected.status_code == 422
    formula = client.patch(f"/api/v1/batch/{batch_id}/rows/2", headers=A, json={"part_name": "=CMD()"})
    assert formula.status_code == 422
