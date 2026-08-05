from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path, PurePosixPath
from zipfile import BadZipFile, ZipFile
from struct import unpack_from

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.exceptions import AppError
from app.models import ExcelBatch, ExcelBatchJob, ExcelBatchRow, Machine
from app.schemas.batch import BatchCartRequest, BatchRowUpdate, BatchTicketRequest
from app.schemas.cart import CartAdd
from app.schemas.tickets import TicketCreate
from app.services.cart import CartOwner, CartService
from app.services.config_service import ConfigService
from app.services.part_search import PartSearchService
from app.services.tickets import TicketService

HEADERS = ["设备类型", "设备品牌", "整机型号", "设备序列号", "发动机型号", "配件名称", "Part Number", "OEM 编号",
           "替代编号", "配件系统", "所需数量", "备注"]
FIELDS = ["machine_type", "machine_brand", "machine_model", "serial_no", "engine_model", "part_name", "part_no",
          "oem_no", "alternate_no", "part_system", "quantity", "note"]
HEADER_ALIASES = {
    "设备类型": "machine_type", "machine type": "machine_type", "equipment type": "machine_type",
    "设备品牌": "machine_brand", "品牌": "machine_brand", "machine brand": "machine_brand", "equipment brand": "machine_brand",
    "整机型号": "machine_model", "设备型号": "machine_model", "machine model": "machine_model", "model": "machine_model",
    "设备序列号": "serial_no", "序列号": "serial_no", "serial number": "serial_no", "serial no": "serial_no",
    "发动机型号": "engine_model", "engine model": "engine_model",
    "配件名称": "part_name", "名称": "part_name", "part name": "part_name", "part description": "part_name",
    "part number": "part_no", "part no": "part_no", "part no.": "part_no", "配件编号": "part_no", "零件号": "part_no",
    "oem 编号": "oem_no", "oem编号": "oem_no", "oem number": "oem_no", "oem no": "oem_no",
    "替代编号": "alternate_no", "替代号": "alternate_no", "alternate number": "alternate_no", "replacement number": "alternate_no",
    "配件系统": "part_system", "系统": "part_system", "part system": "part_system", "system": "part_system",
    "所需数量": "quantity", "数量": "quantity", "quantity": "quantity", "qty": "quantity",
    "备注": "note", "note": "note", "notes": "note", "remark": "note", "remarks": "note",
    "loại thiết bị": "machine_type", "hãng thiết bị": "machine_brand", "mẫu máy": "machine_model",
    "số sê-ri thiết bị": "serial_no", "mẫu động cơ": "engine_model", "tên phụ tùng": "part_name",
    "mã phụ tùng": "part_no", "mã oem": "oem_no", "mã thay thế": "alternate_no",
    "hệ thống phụ tùng": "part_system", "số lượng": "quantity", "ghi chú": "note",
}
IDENTIFIERS = {"part_no", "oem_no", "alternate_no", "machine_model", "serial_no", "engine_model", "part_name"}
XLSX_MIMES = {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/octet-stream"}
XLS_MIMES = {"application/vnd.ms-excel", "application/xls", "application/x-xls", "application/octet-stream"}
OLE_MAGIC = bytes.fromhex("D0CF11E0A1B11AE1")


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


TEMPLATE_HEADERS = {
    "zh": HEADERS,
    "en": ["Machine Type", "Machine Brand", "Machine Model", "Serial Number", "Engine Model", "Part Name",
           "Part Number", "OEM Number", "Alternate Number", "Part System", "Quantity", "Notes"],
    "vi": ["Loại thiết bị", "Hãng thiết bị", "Mẫu máy", "Số sê-ri thiết bị", "Mẫu động cơ", "Tên phụ tùng",
           "Mã phụ tùng", "Mã OEM", "Mã thay thế", "Hệ thống phụ tùng", "Số lượng", "Ghi chú"],
}
TEMPLATE_TEXT = {
    "zh": {"title": "批量匹配", "help": "填写说明", "optional": "可选；有助于提高匹配准确率。",
           "quantity": "必填：正整数。", "part_no": "识别字段之一；建议保留连字符。", "oem_no": "识别字段之一。",
           "model": "识别字段之一；建议同时填写设备品牌。", "name": "识别字段之一；可使用中文、英文或越南语名称。",
           "columns": ["说明", "内容"], "rows": [["必填", "所需数量必须为正整数。"],
             ["识别字段", "每行至少填写配件编号、OEM 编号、替代编号、整机型号、设备序列号、发动机型号或配件名称之一。"],
             ["安全", "不要粘贴公式、宏或外部链接；只填写普通文本和数字。"], ["行数", "最多 500 条数据行；空行自动忽略。"]],
           "example": ["叉车", "Toyota", "8FD30", "", "", "空气滤芯", "12345-67890", "", "", "", 2, ""]},
    "en": {"title": "Batch Match", "help": "Instructions", "optional": "Optional; helps improve matching accuracy.",
           "quantity": "Required: a positive integer.", "part_no": "Identification field; keep hyphens when possible.",
           "oem_no": "Identification field.", "model": "Identification field; also provide the machine brand.",
           "name": "Identification field; Chinese, English, and Vietnamese names are accepted.",
           "columns": ["Item", "Description"], "rows": [["Required", "Quantity must be a positive integer."],
             ["Identification", "Provide at least one identifier or part name on every row."],
             ["Security", "Do not paste formulas, macros, or external links; use plain text and numbers only."],
             ["Rows", "Up to 500 data rows; blank rows are ignored."]],
           "example": ["Forklift", "Toyota", "8FD30", "", "", "Air Filter", "12345-67890", "", "", "", 2, ""]},
    "vi": {"title": "Đối chiếu hàng loạt", "help": "Hướng dẫn", "optional": "Không bắt buộc; giúp tăng độ chính xác.",
           "quantity": "Bắt buộc: số nguyên dương.", "part_no": "Trường nhận dạng; nên giữ dấu gạch nối.",
           "oem_no": "Trường nhận dạng.", "model": "Trường nhận dạng; nên điền cả hãng thiết bị.",
           "name": "Trường nhận dạng; chấp nhận tên tiếng Trung, Anh hoặc Việt.",
           "columns": ["Mục", "Nội dung"], "rows": [["Bắt buộc", "Số lượng phải là số nguyên dương."],
             ["Nhận dạng", "Mỗi dòng phải có ít nhất một mã nhận dạng hoặc tên phụ tùng."],
             ["An toàn", "Không dán công thức, macro hoặc liên kết ngoài; chỉ dùng văn bản và số."],
             ["Số dòng", "Tối đa 500 dòng dữ liệu; dòng trống sẽ bị bỏ qua."]],
           "example": ["Xe nâng", "Toyota", "8FD30", "", "", "Lọc gió", "12345-67890", "", "", "", 2, ""]},
}


def create_template(example: dict | None = None, *, lang: str = "zh") -> bytes:
    wb = Workbook()
    ws = wb.active
    lang = lang if lang in TEMPLATE_HEADERS else "zh"
    labels, copy = TEMPLATE_HEADERS[lang], TEMPLATE_TEXT[lang]
    ws.title = copy["title"]
    example = example if isinstance(example, dict) else {}
    values = list(copy["example"])
    for index, field in enumerate(FIELDS):
        if field in example:
            values[index] = example[field]
        elif labels[index] in example:  # backwards-compatible configured display keys
            values[index] = example[labels[index]]
    instruction_by_field = {"quantity": copy["quantity"], "part_no": copy["part_no"], "oem_no": copy["oem_no"],
                            "machine_model": copy["model"], "part_name": copy["name"]}
    for col, (header, field) in enumerate(zip(labels, FIELDS), 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
        cell.comment = Comment(instruction_by_field.get(field, copy["optional"]), "PartsMatch")
        ws.cell(2, col, values[col - 1])
        ws.column_dimensions[cell.column_letter].width = max(14, len(header) * 2 + 2)
    ws.freeze_panes = "A2"
    help_sheet = wb.create_sheet(copy["help"])
    help_sheet.append(copy["columns"])
    for row in copy["rows"]:
        help_sheet.append(row)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def validate_excel_file(name: str, mime: str | None, content: bytes, config: ConfigService) -> str:
    extension = Path(name).suffix.lower()
    max_bytes = int(config.get("batch.max_file_bytes", 5 * 1024 * 1024))
    if len(content) > min(max(max_bytes, 1), 5 * 1024 * 1024):
        raise AppError("Excel file exceeds the 5MB limit", code="BATCH_FILE_TOO_LARGE", status_code=413)
    if extension == ".xlsx":
        if mime not in XLSX_MIMES or not content.startswith(b"PK\x03\x04"):
            raise AppError("XLSX extension, MIME and content do not match", code="INVALID_EXCEL_FORMAT", status_code=415)
        _validate_zip(content, config)
    elif extension == ".xls":
        if mime not in XLS_MIMES or not content.startswith(OLE_MAGIC):
            raise AppError("XLS extension, MIME and content do not match", code="INVALID_EXCEL_FORMAT", status_code=415)
    else:
        raise AppError("only .xlsx and .xls files are supported", code="INVALID_EXCEL_FORMAT", status_code=415)
    return extension


def _validate_zip(content: bytes, config: ConfigService) -> None:
    max_expanded = min(int(config.get("batch.max_uncompressed_bytes", 20 * 1024 * 1024)), 50 * 1024 * 1024)
    max_ratio = min(float(config.get("batch.max_zip_ratio", 100)), 200.0)
    try:
        with ZipFile(BytesIO(content)) as archive:
            total_compressed = total_expanded = 0
            for item in archive.infolist():
                path = PurePosixPath(item.filename)
                if path.is_absolute() or ".." in path.parts or item.flag_bits & 0x1:
                    raise AppError("unsafe XLSX archive", code="UNSAFE_EXCEL", status_code=400)
                total_compressed += item.compress_size
                total_expanded += item.file_size
                if total_expanded > max_expanded:
                    raise AppError("XLSX expanded content is too large", code="EXCEL_ZIP_BOMB", status_code=413)
            if total_expanded / max(total_compressed, 1) > max_ratio:
                raise AppError("XLSX compression ratio is unsafe", code="EXCEL_ZIP_BOMB", status_code=413)
    except BadZipFile as error:
        raise AppError("invalid XLSX archive", code="INVALID_EXCEL_FORMAT", status_code=415) from error


def parse_excel(extension: str, content: bytes, max_rows: int) -> tuple[list[dict], list[dict]]:
    rows = _xlsx_rows(content) if extension == ".xlsx" else _xls_rows(content)
    header_at = mapping = None
    for index, values in enumerate(rows[:10]):
        candidate = {col: HEADER_ALIASES.get(_text(value).casefold()) for col, value in enumerate(values)}
        candidate = {col: field for col, field in candidate.items() if field}
        if "quantity" in candidate.values() and set(candidate.values()) & IDENTIFIERS:
            header_at, mapping = index, candidate
            break
    if header_at is None or mapping is None:
        raise AppError("header row is missing required quantity and identification columns", code="INVALID_EXCEL_HEADER", status_code=400)
    parsed, errors = [], []
    for physical_index, values in enumerate(rows[header_at + 1:], header_at + 2):
        raw = {_text(rows[header_at][col]): _text(values[col]) if col < len(values) else "" for col in mapping}
        normalized = {field: _text(values[col]) if col < len(values) else "" for col, field in mapping.items()}
        if not any(normalized.values()):
            continue
        if len(parsed) >= min(max(max_rows, 1), 500):
            raise AppError("Excel contains more than 500 data rows", code="BATCH_TOO_MANY_ROWS", status_code=400)
        row_errors = []
        if not any(normalized.get(field, "") for field in IDENTIFIERS):
            row_errors.append("至少填写一个有效识别字段")
        quantity_raw = normalized.get("quantity", "")
        try:
            quantity_number = float(quantity_raw)
            quantity = int(quantity_number)
            if quantity_number != quantity or quantity <= 0:
                raise ValueError
        except (TypeError, ValueError):
            quantity = None
            row_errors.append("所需数量必须为正整数")
        for field, value in normalized.items():
            if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
                row_errors.append(f"{field} 不允许公式或公式样式内容")
        item = {"row_index": physical_index, "raw": raw, "normalized": normalized,
                "quantity": quantity, "errors": list(dict.fromkeys(row_errors))}
        parsed.append(item)
        if item["errors"]:
            errors.append({"row_index": physical_index, "errors": item["errors"]})
    return parsed, errors


def _xlsx_rows(content: bytes) -> list[list]:
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=False, keep_links=False)
        ws = wb.worksheets[0]
        result = []
        for row in ws.iter_rows():
            if any(cell.data_type == "f" for cell in row):
                raise AppError("formulas are not allowed in Excel uploads", code="EXCEL_FORMULA_REJECTED", status_code=400)
            result.append([cell.value for cell in row])
        wb.close()
        return result
    except AppError:
        raise
    except Exception as error:
        raise AppError("unable to parse XLSX file", code="INVALID_EXCEL_FORMAT", status_code=400) from error


def _xls_rows(content: bytes) -> list[list]:
    try:
        from xlrd.compdoc import CompDoc
        document = CompDoc(content, ignore_workbook_corruption=False)
        stream, offset, length = document.locate_named_stream("Workbook")
        if stream is None:
            stream, offset, length = document.locate_named_stream("Book")
        position, end = offset, offset + length
        while stream is not None and position + 4 <= end:
            record_id, record_length = unpack_from("<HH", stream, position)
            if record_id == 0x0006:
                raise AppError("formulas are not allowed in Excel uploads", code="EXCEL_FORMULA_REJECTED", status_code=400)
            position += 4 + record_length
        book = xlrd.open_workbook(file_contents=content, on_demand=True, ignore_workbook_corruption=False)
        sheet = book.sheet_by_index(0)
        rows = [[sheet.cell_value(row, col) for col in range(sheet.ncols)] for row in range(sheet.nrows)]
        book.release_resources()
        return rows
    except AppError:
        raise
    except Exception as error:
        raise AppError("unable to parse XLS file", code="INVALID_EXCEL_FORMAT", status_code=400) from error


class ExcelBatchService:
    def __init__(self, db: Session, owner: CartOwner) -> None:
        self.db, self.owner = db, owner

    def owned(self, batch_id: str, *, lock: bool = False) -> ExcelBatch:
        query = select(ExcelBatch).where(ExcelBatch.id == batch_id, ExcelBatch.owner_key == self.owner.owner_key)
        batch = self.db.scalar(query.with_for_update() if lock else query)
        if batch is None:
            raise AppError("batch not found", code="BATCH_NOT_FOUND", status_code=404)
        return batch

    def rows(self, batch_id: str, *, valid_only: bool = False) -> list[ExcelBatchRow]:
        query = select(ExcelBatchRow).where(ExcelBatchRow.batch_id == batch_id)
        rows = list(self.db.scalars(query.order_by(ExcelBatchRow.row_index)))
        return [row for row in rows if not row.validation_errors] if valid_only else rows

    def match_all(self, batch: ExcelBatch, *, job: ExcelBatchJob | None = None) -> list[dict]:
        batch.status = "matching"
        self.db.commit()
        output = []
        for row in self.rows(batch.id, valid_only=True):
            self._match_row(row)
            if job:
                job.processed_rows += 1
            self.db.commit()
            output.append(self.serialize_row(row))
        batch.status = "matched"
        self.db.commit()
        return output

    def _match_row(self, row: ExcelBatchRow) -> None:
        data, service = row.normalized_content, PartSearchService(self.db)
        if data.get("part_no"):
            result = service.part_number(data["part_no"])
            priority = "Part Number"
        elif data.get("oem_no"):
            result = service.oem(data["oem_no"])
            priority = "OEM"
        elif data.get("alternate_no"):
            result = service.part_number(data["alternate_no"])
            priority = "替代编号"
        elif data.get("machine_model"):
            brand = data.get("machine_brand")
            if not brand:
                brands = list(self.db.scalars(select(Machine.brand).where(Machine.model.ilike(data["machine_model"])).distinct().limit(2)))
                brand = brands[0] if len(brands) == 1 else ""
            result = service.machine(brand, data["machine_model"]) if brand else service.text(data["machine_model"], "en")
            priority = "整机型号"
        elif data.get("engine_model"):
            result = service.engine(data["engine_model"])
            priority = "发动机型号"
        elif data.get("part_name"):
            lang = "zh" if any("\u4e00" <= char <= "\u9fff" for char in data["part_name"]) else "en"
            result = service.text(data["part_name"], lang)
            priority = "配件名称"
        else:
            row.match_status, row.candidates, row.confidence = "insufficient", [], None
            row.match_reason, row.suggested_action = "缺少可匹配识别字段", "supplement"
            return
        candidates = [candidate.model_dump(mode="json") for candidate in result.candidates]
        needs_serial = any(item.get("requires_serial_confirmation") for item in candidates) and not data.get("serial_no")
        status = result.match_status
        if needs_serial:
            status = "need_manual"
        elif not candidates:
            status = "insufficient" if status == "insufficient" else "not_found"
        elif len(candidates) == 1 and status == "exact":
            status = "exact"
        else:
            status = "multiple"
        row.match_status, row.candidates = status, candidates
        row.confidence = max((item.get("confidence", 0) for item in candidates), default=None)
        row.match_reason = f"按 {priority} 优先级匹配；" + (candidates[0].get("reason", "") if candidates else "未找到候选")
        row.suggested_action = {"exact": "confirm", "multiple": "select", "insufficient": "supplement",
                                "not_found": "manual", "need_manual": "manual"}[status]

    def update_row(self, batch: ExcelBatch, row_index: int, payload: BatchRowUpdate) -> dict:
        row = self.db.scalar(select(ExcelBatchRow).where(
            ExcelBatchRow.batch_id == batch.id, ExcelBatchRow.row_index == row_index).with_for_update())
        if row is None:
            raise AppError("batch row not found", code="BATCH_ROW_NOT_FOUND", status_code=404)
        field_map = {"model": "machine_model", "system": "part_system"}
        normalized = dict(row.normalized_content or {})
        raw = dict(row.raw_content or {})
        for source, value in payload.model_dump(exclude_unset=True).items():
            target = field_map.get(source, source)
            normalized[target] = str(value).strip() if isinstance(value, str) else str(value)
            raw[source] = value
            if source == "quantity":
                row.quantity = value
        errors = []
        if not any(normalized.get(field, "") for field in IDENTIFIERS):
            errors.append("至少填写一个有效识别字段")
        if row.quantity is None or row.quantity <= 0:
            errors.append("所需数量必须为正整数")
        row.normalized_content, row.raw_content, row.validation_errors = normalized, raw, errors
        row.match_status, row.candidates, row.confidence = None, [], None
        row.match_reason, row.suggested_action = None, None
        if not errors:
            self._match_row(row)
        batch.valid_rows = sum(not item.validation_errors for item in self.rows(batch.id))
        self.db.commit()
        self.db.refresh(row)
        return self.serialize_row(row)

    @staticmethod
    def serialize_row(row: ExcelBatchRow) -> dict:
        return {"row_index": row.row_index, "raw_content": row.raw_content, "match_status": row.match_status,
                "normalized_content": row.normalized_content, "quantity": row.quantity,
                "candidates": row.candidates or [], "confidence": float(row.confidence) if row.confidence is not None else None,
                "match_reason": row.match_reason, "suggested_action": row.suggested_action,
                "validation_errors": row.validation_errors or [], "ticket_id": row.ticket_id}

    def add_to_cart(self, batch: ExcelBatch, payload: BatchCartRequest) -> dict:
        rows = {row.row_index: row for row in self.rows(batch.id)}
        failures = []
        for selection in payload.selections:
            row = rows.get(selection.row_index)
            candidate_ids = {item.get("part", {}).get("id") for item in (row.candidates if row else [])}
            if row is None:
                failures.append({"row_index": selection.row_index, "error": "row not found"})
            elif row.match_status == "exact" and selection.part_id not in candidate_ids:
                failures.append({"row_index": selection.row_index, "error": "part is not the exact match"})
            elif row.match_status == "multiple" and (not selection.confirmed or selection.part_id not in candidate_ids):
                failures.append({"row_index": selection.row_index, "error": "multiple match requires an explicit confirmed candidate"})
            elif row.match_status not in {"exact", "multiple"}:
                failures.append({"row_index": selection.row_index, "error": "row is not eligible for cart"})
        if failures:
            raise AppError("no cart items were added; fix all selection errors and retry", code="BATCH_CART_VALIDATION_FAILED",
                           status_code=409, data={"atomic": True, "errors": failures})
        cart, added = CartService(self.db, self.owner), []
        try:
            for selection in payload.selections:
                row = rows[selection.row_index]
                item = cart.add(CartAdd(part_id=selection.part_id, quantity=selection.quantity, match_status="exact",
                                        confidence=float(row.confidence) if row.confidence is not None else None,
                                        source="batch"), commit=False)
                added.append({"row_index": row.row_index, "cart_item_id": item.id, "part_id": item.part_id,
                              "quantity": item.quantity})
            self.db.commit()
        except Exception:
            self.db.rollback()
            raise
        return {"atomic": True, "added": added, "errors": []}

    def create_tickets(self, batch: ExcelBatch, payload: BatchTicketRequest) -> dict:
        requested = set(payload.row_indexes) if payload.row_indexes is not None else None
        query = select(ExcelBatchRow).where(ExcelBatchRow.batch_id == batch.id)
        if requested is not None:
            query = query.where(ExcelBatchRow.row_index.in_(requested))
        rows = list(self.db.scalars(query.order_by(ExcelBatchRow.row_index).with_for_update()))
        by_index = {row.row_index for row in rows}
        errors = ([{"row_index": index, "error": "row not found"} for index in sorted(requested - by_index)]
                  if requested is not None else [])
        created, existing = [], []
        ticket_service = TicketService(self.db)
        for row in rows:
            if row.match_status not in {"not_found", "need_manual"}:
                errors.append({"row_index": row.row_index, "error": "row is not eligible for a manual ticket"})
                continue
            if row.ticket_id:
                existing.append({"row_index": row.row_index, "ticket_id": row.ticket_id})
                continue
            data = row.normalized_content
            ticket = ticket_service.create(TicketCreate(
                contact_name=payload.contact_name, contact_info=payload.contact_info, country=payload.country,
                communication_tool=payload.communication_tool, machine_type=data.get("machine_type", ""),
                machine_brand=data.get("machine_brand", ""), machine_model=data.get("machine_model", ""),
                serial_no=data.get("serial_no", ""), engine_model=data.get("engine_model", ""),
                part_description=data.get("part_name") or data.get("part_no") or data.get("oem_no") or "Excel 未匹配项",
                quantity=row.quantity or 1, excel_batch_id=batch.id,
                note=f"Excel 文件: {batch.original_name}; 原始行 {row.row_index}: {row.raw_content}",
            ), self.owner, commit=False)
            row.ticket_id = ticket.id
            self.db.commit()
            created.append({"row_index": row.row_index, "ticket_id": ticket.id})
        return {"created": created, "existing": existing, "errors": errors, "partial_success": bool(errors)}


def run_match_job(bind, job_id: str) -> None:
    from sqlalchemy.orm import Session as SqlSession
    with SqlSession(bind=bind, expire_on_commit=False) as db:
        job = db.get(ExcelBatchJob, job_id)
        if job is None or job.status not in {"queued", "retrying"}:
            return
        job.status, job.attempts, job.started_at, job.error = "running", job.attempts + 1, datetime.now(UTC), None
        db.commit()
        try:
            batch = db.get(ExcelBatch, job.batch_id)
            owner = CartOwner(owner_key=job.owner_key, session_id=batch.session_id, user_id=batch.user_id)
            ExcelBatchService(db, owner).match_all(batch, job=job)
            job.status, job.finished_at = "completed", datetime.now(UTC)
            db.commit()
        except Exception as error:
            db.rollback()
            job = db.get(ExcelBatchJob, job_id)
            job.status, job.error, job.finished_at = "failed", f"{type(error).__name__}: {error}"[:4000], datetime.now(UTC)
            batch = db.get(ExcelBatch, job.batch_id)
            if batch:
                batch.status = "failed"
            db.commit()
